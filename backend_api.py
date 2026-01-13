"""
測試管理系統 - 後端 API v3.0
支援真實檔案上傳和自動嵌入Excel報告
"""

from flask import Flask, request, jsonify, send_file
from flask_cors import CORS
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.drawing.image import Image
from openpyxl.worksheet.datavalidation import DataValidation
import io
import os
import json
from datetime import datetime
from werkzeug.utils import secure_filename

app = Flask(__name__)
CORS(app)  # 允許跨域請求

UPLOAD_FOLDER = 'uploads'
OUTPUT_FOLDER = 'outputs'
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
os.makedirs(OUTPUT_FOLDER, exist_ok=True)

# 允許的圖片格式
ALLOWED_IMAGE_EXTENSIONS = {'png', 'jpg', 'jpeg', 'gif', 'bmp'}

def allowed_file(filename, allowed_extensions=None):
    if allowed_extensions is None:
        return True
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in allowed_extensions

def create_excel_report(data):
    """根據測試資料生成Excel報告,嵌入實際截圖"""
    
    wb = Workbook()
    ws = wb.active
    ws.title = "測試報告"
    
    # 設定欄寬
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 22
    ws.column_dimensions['C'].width = 30
    ws.column_dimensions['D'].width = 25
    ws.column_dimensions['E'].width = 25
    ws.column_dimensions['F'].width = 20
    ws.column_dimensions['G'].width = 25
    ws.column_dimensions['H'].width = 10
    ws.column_dimensions['I'].width = 12
    ws.column_dimensions['J'].width = 12
    ws.column_dimensions['K'].width = 20
    
    # 樣式定義
    title_font = Font(name='微軟正黑體', size=16, bold=True, color='FFFFFF')
    title_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    title_alignment = Alignment(horizontal='center', vertical='center')
    
    header_font = Font(name='微軟正黑體', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='5B9BD5', end_color='5B9BD5', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # 報告標題
    ws.merge_cells('A1:K1')
    ws['A1'] = '測試執行報告'
    ws['A1'].font = title_font
    ws['A1'].fill = title_fill
    ws['A1'].alignment = title_alignment
    ws.row_dimensions[1].height = 30
    
    # 報告資訊
    project_info = data.get('projectInfo', {})
    info_data = [
        ('專案名稱:', project_info.get('projectName', '未設定')),
        ('測試負責人:', project_info.get('testLeader', '未設定')),
        ('測試日期:', project_info.get('testDate', datetime.now().strftime('%Y-%m-%d'))),
        ('測試環境:', project_info.get('testEnv', '未設定'))
    ]
    
    for i, (label, value) in enumerate(info_data, start=2):
        ws[f'A{i}'] = label
        ws[f'A{i}'].font = Font(name='微軟正黑體', bold=True)
        ws[f'A{i}'].alignment = Alignment(horizontal='right')
        ws.merge_cells(f'B{i}:C{i}')
        ws[f'B{i}'] = value
        ws[f'B{i}'].font = Font(name='微軟正黑體')
        ws.row_dimensions[i].height = 20
    
    # 測試案例表頭
    current_row = 7
    headers = ['編號', '測試需求', '測試個案', '前置條件', '測試步驟', 
               '測試資料', '預期結果', '優先級', '執行結果', '測試人員', '執行截圖']
    
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=current_row, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    ws.row_dimensions[current_row].height = 30
    
    # 設定下拉選單
    priority_dv = DataValidation(type="list", formula1='"高,中,低"', allow_blank=False)
    ws.add_data_validation(priority_dv)
    
    result_dv = DataValidation(type="list", formula1='"通過,失敗,待測試,跳過"', allow_blank=False)
    ws.add_data_validation(result_dv)
    
    # 填入測試案例
    current_row += 1
    data_alignment = Alignment(vertical='center', wrap_text=True)
    
    test_cases = data.get('testCases', [])
    
    for test in test_cases:
        row = current_row
        ws.row_dimensions[row].height = 90
        
        # 填入資料
        ws[f'A{row}'] = test['id']
        ws[f'B{row}'] = test['requirement']
        ws[f'C{row}'] = test['case']
        ws[f'D{row}'] = test['precondition']
        ws[f'E{row}'] = test['steps']
        
        # 處理測試資料 - 如果有上傳檔案則顯示超連結
        dataFiles = test.get('dataFiles', [])
        if dataFiles:
            datafiles_text = ', '.join([f'檔案{i+1}' for i in range(len(dataFiles))])
            ws[f'F{row}'] = datafiles_text
            cell = ws[f'F{row}']
            sheet_name = f"{test['id']}_測試資料"
            cell.hyperlink = f"#'{sheet_name}'!A1"
            cell.font = Font(name='微軟正黑體', size=10, color='0563C1', underline='single')
        elif test.get('testData'):
            # 如果沒有上傳檔案但有文字測試資料
            ws[f'F{row}'] = test['testData']
        else:
            ws[f'F{row}'] = ''
        
        # 預期結果
        ws[f'G{row}'] = test.get('expectedResult', '')
        
        ws[f'H{row}'] = test['priority']
        ws[f'I{row}'] = test['result']
        ws[f'J{row}'] = test['tester']
        
        # 處理截圖
        screenshots = test.get('screenshots', [])
        if screenshots:
            screenshots_text = ', '.join([f'截圖{i+1}' for i in range(len(screenshots))])
            ws[f'K{row}'] = screenshots_text
            
            cell = ws[f'K{row}']
            sheet_name = f"{test['id']}_截圖"
            cell.hyperlink = f"#'{sheet_name}'!A1"
            cell.font = Font(name='微軟正黑體', size=10, color='0563C1', underline='single')
        else:
            ws[f'K{row}'] = '無截圖'
        
        # 加入下拉選單
        priority_dv.add(f'H{row}')
        result_dv.add(f'I{row}')
        
        # 套用格式
        for col in range(1, 12):
            cell = ws.cell(row=row, column=col)
            cell.alignment = data_alignment
            cell.border = thin_border
            cell.font = Font(name='微軟正黑體', size=10)
            
            # 執行結果顏色標記
            if col == 9:
                if test['result'] == '通過':
                    cell.fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
                    cell.font = Font(name='微軟正黑體', size=10, color='006100', bold=True)
                elif test['result'] == '失敗':
                    cell.fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
                    cell.font = Font(name='微軟正黑體', size=10, color='9C0006', bold=True)
                elif test['result'] == '待測試':
                    cell.fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
                    cell.font = Font(name='微軟正黑體', size=10, color='9C6500', bold=True)
            
            # 優先級顏色標記
            if col == 8:
                if test['priority'] == '高':
                    cell.font = Font(name='微軟正黑體', size=10, color='C00000', bold=True)
                elif test['priority'] == '中':
                    cell.font = Font(name='微軟正黑體', size=10, color='FF6600', bold=True)
        
        current_row += 1
    
    # 測試統計
    current_row += 2
    ws.merge_cells(f'A{current_row}:K{current_row}')
    ws[f'A{current_row}'] = '測試統計摘要'
    ws[f'A{current_row}'].font = Font(name='微軟正黑體', size=12, bold=True, color='FFFFFF')
    ws[f'A{current_row}'].fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    ws[f'A{current_row}'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[current_row].height = 25
    
    # 計算統計數據
    total = len(test_cases)
    passed = len([tc for tc in test_cases if tc['result'] == '通過'])
    failed = len([tc for tc in test_cases if tc['result'] == '失敗'])
    pending = len([tc for tc in test_cases if tc['result'] == '待測試'])
    pass_rate = passed / total if total > 0 else 0
    
    current_row += 1
    stats = [
        ('總測試案例數', total, '項'),
        ('通過案例數', passed, '項'),
        ('失敗案例數', failed, '項'),
        ('待測試案例數', pending, '項'),
        ('通過率', pass_rate, '%')
    ]
    
    for label, value, unit in stats:
        ws[f'B{current_row}'] = label
        ws[f'B{current_row}'].font = Font(name='微軟正黑體', bold=True)
        ws[f'B{current_row}'].alignment = Alignment(horizontal='right')
        
        ws[f'C{current_row}'] = value
        ws[f'C{current_row}'].font = Font(name='微軟正黑體', size=12, bold=True, color='4472C4')
        ws[f'C{current_row}'].alignment = Alignment(horizontal='center')
        
        if unit == '%':
            ws[f'C{current_row}'].number_format = '0%'
        
        ws[f'D{current_row}'] = unit
        ws[f'D{current_row}'].font = Font(name='微軟正黑體')
        
        current_row += 1
    
    # 為每個有截圖的測試案例創建截圖工作表
    for test in test_cases:
        screenshots = test.get('screenshots', [])
        if screenshots:
            sheet_name = f"{test['id']}_截圖"
            screenshot_ws = wb.create_sheet(sheet_name)
            
            # 標題
            screenshot_ws.merge_cells('A1:D1')
            screenshot_ws['A1'] = f"{test['id']} - {test['case']} 執行截圖"
            screenshot_ws['A1'].font = Font(name='微軟正黑體', size=14, bold=True, color='FFFFFF')
            screenshot_ws['A1'].fill = PatternFill(start_color='5B9BD5', end_color='5B9BD5', fill_type='solid')
            screenshot_ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
            screenshot_ws.row_dimensions[1].height = 30
            
            # 返回連結
            screenshot_ws['A2'] = '← 返回測試報告'
            screenshot_ws['A2'].hyperlink = f"#'測試報告'!A1"
            screenshot_ws['A2'].font = Font(name='微軟正黑體', size=11, underline='single', color='0563C1')
            
            # 為每個截圖嵌入實際圖片
            current_img_row = 4
            for i, screenshot in enumerate(screenshots, 1):
                # 創建截圖標題
                screenshot_ws[f'A{current_img_row}'] = f"截圖{i}: {screenshot['name']}"
                screenshot_ws[f'A{current_img_row}'].font = Font(name='微軟正黑體', size=12, bold=True)
                
                current_img_row += 1
                
                # 嵌入實際圖片
                try:
                    img_path = screenshot['path']
                    if os.path.exists(img_path):
                        excel_img = Image(img_path)
                        
                        # 調整圖片大小 (保持比例,寬度最大600px)
                        max_width = 600
                        if excel_img.width > max_width:
                            ratio = max_width / excel_img.width
                            excel_img.width = max_width
                            excel_img.height = int(excel_img.height * ratio)
                        
                        excel_img.anchor = f'A{current_img_row}'
                        screenshot_ws.add_image(excel_img)
                        
                        # 根據圖片高度調整行高
                        row_height = (excel_img.height / 1.33)
                        screenshot_ws.row_dimensions[current_img_row].height = min(row_height, 500)
                        
                        current_img_row += int(row_height / 15) + 2
                    else:
                        screenshot_ws[f'A{current_img_row}'] = f'⚠️ 圖片檔案不存在: {img_path}'
                        screenshot_ws[f'A{current_img_row}'].font = Font(name='微軟正黑體', color='FF0000')
                        current_img_row += 2
                        
                except Exception as e:
                    screenshot_ws[f'A{current_img_row}'] = f'⚠️ 無法載入圖片: {str(e)}'
                    screenshot_ws[f'A{current_img_row}'].font = Font(name='微軟正黑體', color='FF0000')
                    current_img_row += 2
            
            screenshot_ws.column_dimensions['A'].width = 80
    
    # 為每個有測試資料檔案的測試案例創建測試資料工作表
    for test in test_cases:
        dataFiles = test.get('dataFiles', [])
        if dataFiles:
            sheet_name = f"{test['id']}_測試資料"
            data_ws = wb.create_sheet(sheet_name)
            
            # 標題
            data_ws.merge_cells('A1:D1')
            data_ws['A1'] = f"{test['id']} - {test['case']} 測試資料"
            data_ws['A1'].font = Font(name='微軟正黑體', size=14, bold=True, color='FFFFFF')
            data_ws['A1'].fill = PatternFill(start_color='5B9BD5', end_color='5B9BD5', fill_type='solid')
            data_ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
            data_ws.row_dimensions[1].height = 30
            
            # 返回連結
            data_ws['A2'] = '← 返回測試報告'
            data_ws['A2'].hyperlink = f"#'測試報告'!A1"
            data_ws['A2'].font = Font(name='微軟正黑體', size=11, underline='single', color='0563C1')
            
            # 顯示文字測試資料（如果有）
            current_row = 4
            if test.get('testData'):
                data_ws[f'A{current_row}'] = '測試資料內容:'
                data_ws[f'A{current_row}'].font = Font(name='微軟正黑體', size=12, bold=True)
                current_row += 1
                data_ws[f'A{current_row}'] = test['testData']
                data_ws[f'A{current_row}'].font = Font(name='微軟正黑體', size=10)
                data_ws[f'A{current_row}'].alignment = Alignment(wrap_text=True, vertical='top')
                data_ws.row_dimensions[current_row].height = 60
                current_row += 2
            
            # 為每個測試資料檔案創建區塊
            data_ws[f'A{current_row}'] = '測試資料檔案清單:'
            data_ws[f'A{current_row}'].font = Font(name='微軟正黑體', size=12, bold=True)
            current_row += 1
            
            for i, datafile in enumerate(dataFiles, 1):
                # 創建檔案標題
                data_ws[f'A{current_row}'] = f"檔案{i}: {datafile['name']}"
                data_ws[f'A{current_row}'].font = Font(name='微軟正黑體', size=11, bold=True, color='0563C1')
                current_row += 1
                
                # 檢查是否為圖片檔案
                file_path = datafile['path']
                file_ext = os.path.splitext(file_path)[1].lower()
                
                if file_ext in ['.png', '.jpg', '.jpeg', '.gif', '.bmp']:
                    # 如果是圖片，嵌入圖片
                    try:
                        if os.path.exists(file_path):
                            excel_img = Image(file_path)
                            max_width = 600
                            if excel_img.width > max_width:
                                ratio = max_width / excel_img.width
                                excel_img.width = max_width
                                excel_img.height = int(excel_img.height * ratio)
                            
                            excel_img.anchor = f'A{current_row}'
                            data_ws.add_image(excel_img)
                            
                            row_height = (excel_img.height / 1.33)
                            data_ws.row_dimensions[current_row].height = min(row_height, 500)
                            current_row += int(row_height / 15) + 2
                        else:
                            data_ws[f'A{current_row}'] = f'⚠️ 檔案不存在: {file_path}'
                            data_ws[f'A{current_row}'].font = Font(name='微軟正黑體', color='FF0000')
                            current_row += 2
                    except Exception as e:
                        data_ws[f'A{current_row}'] = f'⚠️ 無法載入圖片: {str(e)}'
                        data_ws[f'A{current_row}'].font = Font(name='微軟正黑體', color='FF0000')
                        current_row += 2
                else:
                    # 如果不是圖片，顯示檔案資訊
                    data_ws[f'A{current_row}'] = f'檔案類型: {file_ext}'
                    data_ws[f'A{current_row}'].font = Font(name='微軟正黑體', size=10)
                    current_row += 1
                    
                    data_ws[f'A{current_row}'] = f'檔案位置: {file_path}'
                    data_ws[f'A{current_row}'].font = Font(name='微軟正黑體', size=9, color='666666')
                    current_row += 1
                    
                    if os.path.exists(file_path):
                        file_size = os.path.getsize(file_path)
                        data_ws[f'A{current_row}'] = f'檔案大小: {file_size / 1024:.1f} KB'
                        data_ws[f'A{current_row}'].font = Font(name='微軟正黑體', size=9, color='666666')
                    else:
                        data_ws[f'A{current_row}'] = '⚠️ 檔案不存在'
                        data_ws[f'A{current_row}'].font = Font(name='微軟正黑體', size=9, color='FF0000')
                    current_row += 2
            
            data_ws.column_dimensions['A'].width = 80
    
    return wb

@app.route('/api/generate-report', methods=['POST'])
def generate_report():
    """接收測試資料並生成Excel報告"""
    try:
        data = request.json
        
        # 生成Excel報告
        wb = create_excel_report(data)
        
        # 保存到記憶體
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        # 生成檔案名稱 - 使用專案名稱和日期時間
        project_name = data.get('projectInfo', {}).get('projectName', '測試報告')
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"{project_name}_{timestamp}.xlsx"
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        print(f"生成報告錯誤: {str(e)}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/upload-file', methods=['POST'])
def upload_file():
    """處理檔案上傳"""
    try:
        if 'file' not in request.files:
            return jsonify({'error': '沒有檔案'}), 400
        
        file = request.files['file']
        test_case_id = request.form.get('testCaseId')
        file_type = request.form.get('fileType')
        
        if file.filename == '':
            return jsonify({'error': '檔案名稱為空'}), 400
        
        # 安全的檔案名稱
        filename = secure_filename(file.filename)
        
        # 建立資料夾
        upload_path = os.path.join(UPLOAD_FOLDER, test_case_id, file_type)
        os.makedirs(upload_path, exist_ok=True)
        
        # 儲存檔案
        filepath = os.path.join(upload_path, filename)
        file.save(filepath)
        
        print(f"✅ 檔案已上傳: {filepath}")
        
        return jsonify({
            'success': True,
            'filename': filename,
            'path': filepath
        })
        
    except Exception as e:
        print(f"❌ 上傳檔案錯誤: {str(e)}")
        return jsonify({'error': str(e)}), 500

@app.route('/health', methods=['GET'])
def health():
    """健康檢查端點"""
    return jsonify({
        'status': 'ok',
        'upload_folder': UPLOAD_FOLDER,
        'upload_folder_exists': os.path.exists(UPLOAD_FOLDER)
    })

if __name__ == '__main__':
    print("=" * 60)
    print("🚀 測試管理系統後端 API v3.0 啟動中...")
    print("=" * 60)
    print(f"📍 URL: http://localhost:5000")
    print(f"📁 上傳目錄: {os.path.abspath(UPLOAD_FOLDER)}")
    print("")
    print("📊 可用的 API 端點:")
    print("   ├─ POST /api/generate-report   生成測試報告(嵌入實際截圖)")
    print("   ├─ POST /api/upload-file       上傳檔案到伺服器")
    print("   └─ GET  /health                健康檢查")
    print("")
    print("✨ v3.0 新功能:")
    print("   ✓ 真實檔案上傳到伺服器")
    print("   ✓ 自動嵌入截圖到 Excel")
    print("   ✓ 上傳檔案會儲存在 uploads 目錄")
    print("=" * 60)
    app.run(debug=True, host='0.0.0.0', port=5000)
